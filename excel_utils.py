import os
import pandas as pd
import numpy as np
import xlsxwriter
from openpyxl.utils import get_column_letter

def _write_cell(ws, row, col, value):
    """
    Write a value with automatic type handling that works around
    the numpy-scalar limitation in XlsxWriter.
    """
    if value is None or (isinstance(value, float) and np.isnan(value)):
        ws.write_blank(row, col, None)
    elif isinstance(value, (np.integer, np.floating)):
        ws.write_number(row, col, float(value))
    else:
        ws.write(row, col, value)

def write_metrics_to_excel(
    df_metrics,
    dataset_name,
    k_value,
    excel_path="results/clustering_metrics.xlsx"
):
    # 1) Prepare new block of metrics
    df_new = df_metrics.copy()
    df_new.insert(0, "Dataset", dataset_name)
    df_new.insert(1, "K", k_value)

    # 2) Ensure output directory
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)

    # 3) Read existing “All-Metrics” if present, else start fresh
    if os.path.exists(excel_path):
        df_all = pd.read_excel(excel_path, sheet_name="All-Metrics")
        df_all = pd.concat([df_all, df_new], ignore_index=True)
    else:
        df_all = df_new.copy()

    # 4) Compute row/column counts for formulas
    n_rows = len(df_all) + 1   # +1 for header row
    n_cols = df_all.shape[1]
    last_col = get_column_letter(n_cols)

    # find the column-letter for “Method”
    method_idx = df_all.columns.get_loc("Method")
    method_col = get_column_letter(method_idx + 1)

    # 5) Create a brand‐new XLSX with XlsxWriter
    workbook = xlsxwriter.Workbook(excel_path)
    bold_fmt = workbook.add_format({"bold": True})

    # --- All-Metrics sheet ---
    ws_all = workbook.add_worksheet("All-Metrics")
    # write header
    for col_idx, col in enumerate(df_all.columns):
        ws_all.write(0, col_idx, col, bold_fmt)
    # write data rows
    for row_idx, row in enumerate(df_all.itertuples(index=False, name=None), start=1):
        for col_idx, value in enumerate(row):
            ws_all.write(row_idx, col_idx, value)

    # --- One sheet per Method, with FILTER formula in A2 ---
    for method in df_all["Method"].unique():
        sheet_name = method[:31]  # Excel’s 31-char limit
        ws = workbook.add_worksheet(sheet_name)
        # build absolute ranges
        filter_range   = f"'All-Metrics'!$A$2:${last_col}${n_rows}"
        criteria_range = f"'All-Metrics'!${method_col}$2:${method_col}${n_rows}"
        formula = (
            f'=FILTER({filter_range}, '
            f'{criteria_range}="{method}", '
            '"" )'
        )
        ws.write_formula(1, 0, formula)

    # --- Overview pivot of mean Silhouette ---
    if "Silhouette" in df_all.columns:
        pivot = (
            df_all
              .pivot_table(index="Dataset",
                           columns="Method",
                           values="Silhouette",
                           aggfunc="mean")
              .round(3)
              .reset_index()
        )
        ws_ov = workbook.add_worksheet("Overview")
        # header
        for col_idx, col in enumerate(pivot.columns):
            ws_ov.write(0, col_idx, col, bold_fmt)
        # data
        for row_idx, row in enumerate(pivot.itertuples(index=False, name=None), start=1):
            for col_idx, value in enumerate(row):
                _write_cell(ws_ov, row_idx, col_idx, value)
    workbook.close()
